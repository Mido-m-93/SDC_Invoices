"""月次締めレポート（月初 09:00 JST）。

前月の請求・入金・支払を集計して 32_V_月次ダッシュボード に書き込み、Teamsへ投稿する。
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, datetime
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

from reconcile import _parse_date, _rows, detect_alerts
from teams_notifier import Alert, post_alerts

HERE = Path(__file__).parent
SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
LOG_DIR = HERE / "logs"

logger = logging.getLogger("monthly_report")


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def _prev_month(today: date) -> tuple[date, date]:
    first_this = today.replace(day=1)
    last_prev = first_this.replace(day=1)
    last_prev = (first_this.toordinal() - 1)
    # 簡易: 今月1日 - 1日 = 先月末
    last_prev_date = date.fromordinal(first_this.toordinal() - 1)
    first_prev = last_prev_date.replace(day=1)
    return first_prev, last_prev_date


def build(xlsx: Path, apply_: bool) -> list[Alert]:
    wb = load_workbook(xlsx, data_only=True)
    invoices = _rows(wb["20_T_請求明細"])
    receipts = _rows(wb["21_T_入金明細"])
    payments = _rows(wb["22_T_支払明細"])
    contracts = _rows(wb["12_M_契約"])

    today = date.today()
    start, end = _prev_month(today)
    ym = f"{start.year:04d}-{start.month:02d}"

    def _in_month(v):
        d = _parse_date(v)
        return d is not None and start <= d <= end

    invoiced = sum(float(i.get("amount_incl_tax") or 0) for i in invoices if _in_month(i.get("issue_date")))
    received = sum(float(r.get("amount") or 0) for r in receipts if _in_month(r.get("received_date") or r.get("date")))
    paid = sum(float(p.get("amount") or 0) for p in payments if _in_month(p.get("paid_date") or p.get("date")))

    ar_balance = sum(
        float(i.get("amount_incl_tax") or 0)
        for i in invoices if i.get("status") not in ("paid", "cancelled")
    )
    ap_balance = sum(
        float(p.get("amount") or 0) for p in payments if p.get("status") != "paid"
    )

    new_contracts = sum(1 for c in contracts if _in_month(c.get("contract_start")))
    ended_contracts = sum(1 for c in contracts if _in_month(c.get("contract_end")))

    alerts_all = detect_alerts(xlsx)
    alerts_open = len(alerts_all)

    if apply_:
        ws = wb["32_V_月次ダッシュボード"]
        # 既存の同year_month行を置換
        to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] == ym:
                to_delete.append(idx)
        for idx in reversed(to_delete):
            ws.delete_rows(idx, 1)
        ws.append([
            ym, invoiced, received, paid, ar_balance, ap_balance,
            new_contracts, ended_contracts, alerts_open, 0, "",
        ])
        wb.save(xlsx)
        logger.info("32_V_月次ダッシュボード 更新: %s", ym)

    summary = Alert(
        severity="🟡",
        category="monthly_close",
        title=f"月次締めレポート {ym}",
        message=f"期間: {start} 〜 {end}",
        fields={
            "請求合計": f"¥{invoiced:,.0f}",
            "入金合計": f"¥{received:,.0f}",
            "支払合計": f"¥{paid:,.0f}",
            "売掛残": f"¥{ar_balance:,.0f}",
            "買掛残": f"¥{ap_balance:,.0f}",
            "新規契約": f"{new_contracts}件",
            "終了契約": f"{ended_contracts}件",
            "未解決アラート": f"{alerts_open}件",
        },
    )
    return [summary]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--prod", action="store_true")
    args = parser.parse_args()
    setup_logging(LOG_DIR / f"monthly_report_{datetime.now():%Y-%m-%d}.log")

    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    load_dotenv(HERE / ".env")

    alerts = build(xlsx, apply_=args.apply)
    post_alerts(alerts, prod=args.prod, title="RCP 月次締めレポート")
    return 0


if __name__ == "__main__":
    sys.exit(main())
