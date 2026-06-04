"""契約↔請求↔入金 突合エンジン。

統合管理Excelの M_/T_ シートを読み、31_V_週次チェック にアラートを書き出し、
Teamsへ投稿する。

検出カテゴリ:
    🔴 missing_invoice     : Status=Contracted以降なのにMF請求書なし
    🔴 overdue_payment     : 入金予定+3営業日を超えて未入金
    🔴 overdue_payable     : 支払予定日超過
    🟡 contract_ending     : 契約終了30日前
    🟠 amount_mismatch     : 請求額と契約期待額の乖離 ±5%超
    🟠 unmapped_invoice    : MF側には存在するが案件ひも付けできていない請求書
    🟢 partner_queue       : fuzzy中信頼度で要確認の取引先

実行:
    python reconcile.py                    # 検出のみ、Teamsへはtest投稿
    python reconcile.py --apply            # Excel 31_V_ を更新 + test投稿
    python reconcile.py --apply --prod     # Excel更新 + 本番チャネル投稿
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

from teams_notifier import Alert, post_alerts

HERE = Path(__file__).parent
SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
LOG_DIR = HERE / "logs"

logger = logging.getLogger("reconcile")

AMOUNT_TOLERANCE = 0.05  # 5%
OVERDUE_GRACE_DAYS = 3
CONTRACT_ENDING_DAYS = 30


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def _parse_date(v) -> date | None:
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _rows(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def detect_alerts(xlsx: Path) -> list[Alert]:
    wb = load_workbook(xlsx, data_only=True)
    projects = _rows(wb["11_M_案件"])
    partners = _rows(wb["10_M_取引先"])
    contracts = _rows(wb["12_M_契約"])
    invoices = _rows(wb["20_T_請求明細"])
    receipts = _rows(wb["21_T_入金明細"])
    payments = _rows(wb["22_T_支払明細"])

    today = date.today()
    alerts: list[Alert] = []

    contracts_by_project = {c["project_id"]: c for c in contracts if c.get("project_id")}
    invoices_by_partner: dict[str, list[dict]] = {}
    for inv in invoices:
        invoices_by_partner.setdefault(inv.get("partner_id") or "", []).append(inv)

    # 🔴 missing_invoice
    for p in projects:
        if p.get("status") in ("Contracted", "Delivery", "Delivered"):
            pid = p.get("partner_id")
            invs = invoices_by_partner.get(pid, [])
            if not invs:
                alerts.append(Alert(
                    severity="🔴",
                    category="missing_invoice",
                    title=f"請求書が見つかりません: {p.get('project_name')}",
                    message=f"Status={p.get('status')} ですが、MFに該当取引先の請求書がありません",
                    fields={
                        "project_id": p.get("project_id", ""),
                        "partner_id": pid or "",
                        "status": p.get("status", ""),
                    },
                ))

    # 🔴 overdue_payment（売掛）
    paid_invoice_ids = {r.get("mf_invoice_id") for r in receipts if r.get("mf_invoice_id")}
    for inv in invoices:
        due = _parse_date(inv.get("due_date"))
        if not due:
            continue
        if inv.get("status") == "paid" or inv.get("mf_invoice_id") in paid_invoice_ids:
            continue
        if today > due + timedelta(days=OVERDUE_GRACE_DAYS):
            alerts.append(Alert(
                severity="🔴",
                category="overdue_payment",
                title=f"未入金超過: {inv.get('invoice_number')}",
                message=f"due={due} / 本日={today} / 猶予{OVERDUE_GRACE_DAYS}日超過",
                fields={
                    "mf_invoice_id": inv.get("mf_invoice_id", ""),
                    "partner_id": inv.get("partner_id", ""),
                    "amount": str(inv.get("amount_incl_tax") or ""),
                },
            ))

    # 🔴 overdue_payable（買掛）
    for pay in payments:
        if pay.get("status") == "overdue":
            alerts.append(Alert(
                severity="🔴",
                category="overdue_payable",
                title=f"支払期日超過: {pay.get('memo') or pay.get('mf_transaction_id')}",
                message=f"due={pay.get('due_date')} / 未払",
                fields={
                    "mf_transaction_id": pay.get("mf_transaction_id", ""),
                    "partner_id": pay.get("partner_id", ""),
                    "amount": str(pay.get("amount") or ""),
                },
            ))

    # 🟡 contract_ending
    for c in contracts:
        end = _parse_date(c.get("contract_end"))
        if end and 0 <= (end - today).days <= CONTRACT_ENDING_DAYS:
            alerts.append(Alert(
                severity="🟡",
                category="contract_ending",
                title=f"契約終了 {(end - today).days}日前: {c.get('contract_id')}",
                message=f"project_id={c.get('project_id')} / end={end}",
                fields={"contract_id": c.get("contract_id", "")},
            ))

    # 🟠 amount_mismatch
    for inv in invoices:
        pid = inv.get("partner_id")
        # partner_id から紐付く契約月額を探す（project_idが埋まるまでは部分実装）
        expected = None
        for c in contracts:
            if c.get("partner_id") == pid and c.get("amount_monthly_k_jpy"):
                expected = float(c["amount_monthly_k_jpy"]) * 1000
                break
        actual = inv.get("amount_excl_tax")
        if expected and actual:
            diff_ratio = abs(float(actual) - expected) / expected
            if diff_ratio > AMOUNT_TOLERANCE:
                alerts.append(Alert(
                    severity="🟠",
                    category="amount_mismatch",
                    title=f"請求額乖離: {inv.get('invoice_number')}",
                    message=f"契約月額 {expected:,.0f} vs 請求 {actual:,.0f} ({diff_ratio*100:.1f}%)",
                    fields={"mf_invoice_id": inv.get("mf_invoice_id", "")},
                ))

    # 🟠 unmapped_invoice（project_id 未解決）
    for inv in invoices:
        if not inv.get("project_id"):
            alerts.append(Alert(
                severity="🟠",
                category="unmapped_invoice",
                title=f"案件未紐付け請求書: {inv.get('invoice_number')}",
                message="20_T_請求明細.project_id が空です",
                fields={
                    "mf_invoice_id": inv.get("mf_invoice_id", ""),
                    "partner_id": inv.get("partner_id", ""),
                },
            ))

    # 🟢 partner_queue
    for p in partners:
        if p.get("mf_match_status") == "queued":
            alerts.append(Alert(
                severity="🟢",
                category="partner_queue",
                title=f"取引先マッチング要確認: {p.get('partner_name')}",
                message=f"fuzzy score={p.get('mf_match_score')}",
                fields={
                    "partner_id": p.get("partner_id", ""),
                    "mf_partner_id": p.get("mf_partner_id", ""),
                },
            ))

    logger.info("検出アラート: %d件", len(alerts))
    by_cat: dict[str, int] = {}
    for a in alerts:
        by_cat[a.category] = by_cat.get(a.category, 0) + 1
    for k, v in by_cat.items():
        logger.info("  %s: %d", k, v)
    return alerts


def write_alerts_to_excel(xlsx: Path, alerts: list[Alert]) -> None:
    wb = load_workbook(xlsx)
    ws = wb["31_V_週次チェック"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    now = datetime.now().isoformat(timespec="seconds")
    for a in alerts:
        ws.append([
            now,
            a.severity,
            a.category,
            a.fields.get("project_id", ""),
            a.fields.get("partner_id", ""),
            a.fields.get("mf_invoice_id", ""),
            "",  # expected
            "",  # actual
            a.message,
            a.action_label or "",
            False,
        ])
    wb.save(xlsx)
    logger.info("✅ 31_V_週次チェック 更新完了 %d件", len(alerts))


def run(apply_: bool, prod: bool) -> int:
    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    if not xlsx.exists():
        logger.error("Excel未生成: %s", xlsx)
        return 2

    load_dotenv(HERE / ".env")
    alerts = detect_alerts(xlsx)

    if apply_:
        write_alerts_to_excel(xlsx, alerts)

    post_alerts(alerts, prod=prod, title="RCP自動チェック（日次）")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="契約↔請求↔入金 突合 + Teams通知")
    parser.add_argument("--apply", action="store_true", help="31_V_週次チェック シートを更新")
    parser.add_argument("--prod", action="store_true", help="Teamsを本番チャネルへ投稿")
    args = parser.parse_args()
    setup_logging(LOG_DIR / f"reconcile_{datetime.now():%Y-%m-%d}.log")
    return run(args.apply, args.prod)


if __name__ == "__main__":
    sys.exit(main())
