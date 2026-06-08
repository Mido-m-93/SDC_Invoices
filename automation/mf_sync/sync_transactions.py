"""MF会計取引 → 21_T_入金明細 / 22_T_支払明細 への同期。

type: receipt → 21_T_入金明細
type: payment → 22_T_支払明細
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

HERE = Path(__file__).parent
SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
LOG_DIR = HERE / "logs"
SAMPLES_DIR = HERE / "samples"

logger = logging.getLogger("sync_transactions")


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def load_transactions(mock: bool, since: str, until: str) -> list[dict]:
    if mock:
        return json.loads((SAMPLES_DIR / "mf_transactions_samples.json").read_text(encoding="utf-8"))
    load_dotenv(HERE / ".env")
    from mf_client import MFClient, MFConfig
    client = MFClient(MFConfig.from_env())
    return list(client.list_transactions(since, until))


def _partner_lookup(xlsx_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["10_M_取引先"]
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pid, _, _, mf_id, *_ = row
        if mf_id:
            lookup[str(mf_id)] = pid
    return lookup


def run(mode: str, since: str, until: str) -> int:
    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    if not xlsx.exists():
        logger.error("Excel未生成: %s", xlsx)
        return 2

    txs = load_transactions(mode == "mock", since, until)
    lookup = _partner_lookup(xlsx)
    logger.info("取引 %d件 (%s 〜 %s)", len(txs), since, until)

    if mode == "dry-run":
        (LOG_DIR / "transactions_preview.json").write_text(
            json.dumps(txs, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info("[DRY-RUN] Excelは更新せず")
        return 0

    wb = load_workbook(xlsx)
    receipts = wb["21_T_入金明細"]
    payments = wb["22_T_支払明細"]
    for ws in (receipts, payments):
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    now_iso = datetime.now().isoformat(timespec="seconds")
    r_count = p_count = 0
    for tx in txs:
        rc_pid = lookup.get(str(tx.get("partner_id"))) or ""
        if tx.get("type") == "receipt":
            receipts.append([
                tx.get("id") or "",
                tx.get("invoice_id") or "",
                rc_pid,
                "", "",  # project_id, contract_id
                tx.get("date") or "",
                tx.get("amount") or 0,
                tx.get("currency") or "JPY",
                tx.get("bank_account") or "",
                tx.get("account_item") or "",
                tx.get("memo") or "",
                bool(tx.get("invoice_id")),
                now_iso,
            ])
            r_count += 1
        elif tx.get("type") == "payment":
            payments.append([
                tx.get("id") or "",
                rc_pid,
                "",  # project_id
                tx.get("date") or "",
                tx.get("due_date") or "",
                tx.get("amount") or 0,
                tx.get("currency") or "JPY",
                tx.get("category") or "",
                tx.get("account_item") or "",
                tx.get("memo") or "",
                tx.get("status") or "",
                now_iso,
            ])
            p_count += 1

    wb.save(xlsx)
    logger.info("✅ 21_T_入金明細 %d行 / 22_T_支払明細 %d行", r_count, p_count)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--mock", action="store_true")
    mode.add_argument("--apply", action="store_true")
    parser.add_argument("--since", default="2026-01-01")
    parser.add_argument("--until", default=datetime.now().strftime("%Y-%m-%d"))
    args = parser.parse_args()
    m = "mock" if args.mock else ("apply" if args.apply else "dry-run")
    setup_logging(LOG_DIR / f"sync_transactions_{datetime.now():%Y-%m-%d}.log")
    return run(m, args.since, args.until)


if __name__ == "__main__":
    sys.exit(main())
