"""
Notion v2パイプライン → Excel M_取引先 投入スクリプト。

Notion から Company 一覧（重複排除）を取得し、
RCP_契約請求支払統合管理.xlsx の 10_M_取引先 シートに書き込む。

実行:
    op run --env-file=/tmp/mf_op_env.txt -- python populate_master.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent / "notion_sync"))

from notion_client import NotionClient, NotionConfig  # noqa: E402

SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
V2_DATA_SOURCE_ID = "dbeaedd1-6635-4607-92b1-1ca54c9df7e6"

# 突合対象ステータス（アクティブな案件のみ）
ACTIVE_STATUSES = {"Contracted", "Delivery", "Delivered", "Negotiation", "Proposal", "Closing"}


def fetch_companies() -> list[str]:
    """Notion v2 パイプラインから会社名（重複排除・空除外）を取得。"""
    client = NotionClient(NotionConfig.from_env())
    companies: dict[str, None] = {}  # 順序保持の重複排除
    count = 0
    print("Notion からデータ取得中...", flush=True)
    for page in client.query_data_source(V2_DATA_SOURCE_ID):
        count += 1
        props = page.get("properties", {})
        company = ""
        # Company プロパティ取得
        c_prop = props.get("Company", {})
        if isinstance(c_prop, dict):
            rich = c_prop.get("rich_text", [])
            if rich:
                company = "".join(r.get("plain_text", "") for r in rich).strip()
            elif "title" in c_prop:
                company = "".join(r.get("plain_text", "") for r in c_prop["title"]).strip()

        status = ""
        s_prop = props.get("Status", {})
        if isinstance(s_prop, dict):
            sel = s_prop.get("select")
            if sel:
                status = sel.get("name", "")

        if company and status in ACTIVE_STATUSES:
            companies[company] = None

    print(f"  → {count} ページ取得, {len(companies)} 社（アクティブ）", flush=True)
    return list(companies.keys())


def update_excel(companies: list[str], xlsx_path: Path) -> None:
    """M_取引先 シートのダミーデータを消去してNotion企業で上書き。"""
    wb = load_workbook(xlsx_path)
    ws = wb["10_M_取引先"]

    # ヘッダ行確認
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    print(f"  ヘッダ: {headers}")

    # 既存データを消去（2行目以降）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 新データ書き込み
    for i, company in enumerate(companies, start=2):
        partner_id = f"RC-PRT-{i-1:04d}"
        ws.cell(i, 1).value = partner_id       # partner_id
        ws.cell(i, 2).value = company           # partner_name
        ws.cell(i, 3).value = company           # partner_name_raw
        # mf_partner_id, score, status は sync_partners.py が埋める
        ws.cell(i, 4).value = None
        ws.cell(i, 5).value = None
        ws.cell(i, 6).value = "unmatched"

    wb.save(xlsx_path)
    print(f"  → {len(companies)} 社を M_取引先 に書き込み完了", flush=True)


def main() -> None:
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    if not xlsx.exists():
        print(f"[ERROR] Excel が見つかりません: {xlsx}", file=sys.stderr)
        sys.exit(1)

    companies = fetch_companies()
    if not companies:
        print("[WARN] 取得した会社名がゼロ件でした。Notionトークンと権限を確認してください。")
        sys.exit(1)

    print(f"\n取得会社一覧 ({len(companies)} 社):")
    for c in companies:
        print(f"  - {c}")

    update_excel(companies, xlsx)
    print(f"\n✅ 完了。次: python sync_partners.py --dry-run")


if __name__ == "__main__":
    main()
