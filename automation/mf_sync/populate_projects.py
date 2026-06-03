"""
Notion v2パイプライン → Excel 11_M_案件 投入スクリプト。

Notion から全案件（Company / Project Name / Status / Revenue 等）を取得し、
partner_id (M_取引先) と紐付けて 11_M_案件 シートに書き込む。

実行:
    op run --env-file=/tmp/mf_op_env.txt -- python populate_projects.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent / "notion_sync"))

from notion_client import NotionClient, NotionConfig  # noqa: E402

SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
V2_DATA_SOURCE_ID = "dbeaedd1-6635-4607-92b1-1ca54c9df7e6"

ACTIVE_STATUSES = {"Contracted", "Delivery", "Delivered", "Negotiation", "Proposal", "Closing"}


def _text(prop: dict) -> str:
    if not prop:
        return ""
    rich = prop.get("rich_text", []) or prop.get("title", [])
    return "".join(r.get("plain_text", "") for r in rich).strip()


def _select(prop: dict) -> str:
    if not prop:
        return ""
    sel = prop.get("select")
    return sel.get("name", "") if sel else ""


def _number(prop: dict) -> float | None:
    if not prop:
        return None
    v = prop.get("number")
    return float(v) if v is not None else None


def _date(prop: dict) -> str | None:
    if not prop:
        return None
    d = prop.get("date")
    if d and d.get("start"):
        return d["start"]
    return None


def _url(prop: dict) -> str:
    if not prop:
        return ""
    return prop.get("url") or ""


def fetch_projects() -> list[dict]:
    client = NotionClient(NotionConfig.from_env())
    projects: list[dict] = []
    print("Notion からプロジェクト取得中...", flush=True)
    for page in client.query_data_source(V2_DATA_SOURCE_ID):
        props = page.get("properties", {})
        status = _select(props.get("Status", {}))
        if status not in ACTIVE_STATUSES:
            continue
        company = _text(props.get("Company", {}))
        project_name = _text(props.get("Project Name", {}))
        if not project_name:
            # タイトルを直接取得
            title_prop = props.get("Project Name", {})
            project_name = "".join(
                r.get("plain_text", "") for r in title_prop.get("title", [])
            ).strip()
        projects.append({
            "notion_url": f"https://www.notion.so/{page['id'].replace('-','')}",
            "project_name": project_name or "(no title)",
            "company": company,
            "status": status,
            "source_db": _select(props.get("Source_DB", {})),
            "sales_in_charge": _text(props.get("Sales_In_Charge", {})),
            "region": _select(props.get("Region", {})),
            "project_type": _text(props.get("Project_Type", {})),
            "new_continuing": _select(props.get("New_Continuing", {})),
            "industry": _text(props.get("Industry", {})),
            "probability": _select(props.get("Probability", {})),
            "priority": _select(props.get("Priority", {})),
            "expected_close": _date(props.get("Expected_Close", {})),
            "last_discussion": _date(props.get("Last_Discussion", {})),
            "revenue_k_jpy": _number(props.get("Revenue_千円", {})),
            "revenue_monthly_k_jpy": _number(props.get("Revenue_月_千円", {})),
            "pj_term_months": _number(props.get("PJ_Term_months", {})),
            "next_step": _text(props.get("Next_Step", {})),
            "original_url": _url(props.get("Original_URL", {})),
        })
    print(f"  → {len(projects)} 件取得", flush=True)
    return projects


def load_partner_map(xlsx_path: Path) -> dict[str, str]:
    """M_取引先から company_name → partner_id の辞書を作る。"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["10_M_取引先"]
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            break
        partner_id, partner_name = row[0], row[1]
        if partner_name:
            mapping[partner_name] = partner_id
    return mapping


def update_excel(projects: list[dict], partner_map: dict[str, str], xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb["11_M_案件"]

    # 既存データ消去
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    now_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    for i, p in enumerate(projects, start=2):
        partner_id = partner_map.get(p["company"])  # None if unmatched
        ws.cell(i, 1).value  = f"RC-PRJ-{i-1:04d}"       # project_id
        ws.cell(i, 2).value  = p["project_name"]           # project_name
        ws.cell(i, 3).value  = partner_id                  # partner_id
        ws.cell(i, 4).value  = p["status"]                 # status
        ws.cell(i, 5).value  = p["source_db"]              # source_db
        ws.cell(i, 6).value  = p["sales_in_charge"]        # sales_in_charge
        ws.cell(i, 7).value  = p["region"]                 # region
        ws.cell(i, 8).value  = p["project_type"]           # project_type
        ws.cell(i, 9).value  = p["new_continuing"]         # new_continuing
        ws.cell(i, 10).value = p["industry"]               # industry
        ws.cell(i, 11).value = p["probability"]            # probability
        ws.cell(i, 12).value = None                        # strategic_fit
        ws.cell(i, 13).value = p["priority"]               # priority
        ws.cell(i, 14).value = p["expected_close"]         # expected_close
        ws.cell(i, 15).value = p["last_discussion"]        # last_discussion
        ws.cell(i, 16).value = p["revenue_k_jpy"]          # revenue_k_jpy
        ws.cell(i, 17).value = p["revenue_monthly_k_jpy"]  # revenue_monthly_k_jpy
        ws.cell(i, 18).value = p["pj_term_months"]         # pj_term_months
        ws.cell(i, 19).value = p["next_step"]              # next_step
        ws.cell(i, 20).value = p["notion_url"]             # notion_url
        ws.cell(i, 21).value = p["original_url"]           # original_url
        ws.cell(i, 22).value = now_str                     # last_synced_at

    wb.save(xlsx_path)
    unlinked = sum(1 for p in projects if not partner_map.get(p["company"]))
    print(f"  → {len(projects)} 件書き込み完了（取引先未紐付け: {unlinked} 件）", flush=True)


def main() -> None:
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    if not xlsx.exists():
        print(f"[ERROR] Excel が見つかりません: {xlsx}", file=sys.stderr)
        sys.exit(1)

    partner_map = load_partner_map(xlsx)
    print(f"取引先マップ: {len(partner_map)} 件")

    projects = fetch_projects()
    update_excel(projects, partner_map, xlsx)
    print(f"\n✅ 完了。次: python reconcile.py")


if __name__ == "__main__":
    main()
