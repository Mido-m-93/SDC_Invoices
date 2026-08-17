"""MF請求書 → 20_T_請求明細 への同期。

mockモード:
    samples/mf_invoices_samples.json を読んで Excel に流し込む
本番:
    MFClient経由で /v3/invoice_template/billings を取得（スコープ: read）
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

HERE = Path(__file__).parent
SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
LOG_DIR = HERE / "logs"
SAMPLES_DIR = HERE / "samples"

logger = logging.getLogger("sync_invoices")


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def load_invoices(mock: bool) -> list[dict]:
    if mock:
        return json.loads((SAMPLES_DIR / "mf_invoices_samples.json").read_text(encoding="utf-8"))
    load_dotenv(HERE / ".env")
    from mf_client import MFClient, MFConfig
    client = MFClient(MFConfig.from_env())
    return list(client.list_invoices())


def build_partner_lookup(xlsx_path: Path) -> dict[str, str]:
    """mf_partner_id → RC partner_id"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["10_M_取引先"]
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        partner_id, _, _, mf_id, *_ = row
        if mf_id:
            lookup[str(mf_id)] = partner_id
    return lookup


def run(mode: str) -> int:
    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    if not xlsx.exists():
        logger.error("Excel未生成: %s", xlsx)
        return 2

    invoices = load_invoices(mode == "mock")
    lookup = build_partner_lookup(xlsx)
    logger.info("請求書 %d件, 取引先マップ %d件", len(invoices), len(lookup))

    if mode in ("dry-run", "mock") and mode != "mock":
        preview = [{
            "mf_invoice_id": inv.get("id"),
            "partner_resolved": lookup.get(str(inv.get("partner_id"))),
            **inv,
        } for inv in invoices]
        (LOG_DIR / "invoice_sync_preview.json").write_text(
            json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info("[DRY-RUN] Excelは更新せず")
        return 0

    # mock / apply は Excel書き込み
    wb = load_workbook(xlsx)
    ws = wb["20_T_請求明細"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    now_iso = datetime.now().isoformat(timespec="seconds")
    for inv in invoices:
        ws.append([
            inv.get("id") or "",
            inv.get("invoice_number") or "",
            inv.get("partner_id") or "",
            lookup.get(str(inv.get("partner_id"))) or "",
            "",  # project_id
            "",  # contract_id
            inv.get("issue_date") or "",
            inv.get("due_date") or "",
            inv.get("billing_period_from") or "",
            inv.get("billing_period_to") or "",
            inv.get("amount_excl_tax") or 0,
            inv.get("tax") or 0,
            inv.get("amount_incl_tax") or 0,
            inv.get("currency") or "JPY",
            inv.get("status") or "",
            inv.get("url") or "",
            now_iso,
        ])
    wb.save(xlsx)
    logger.info("✅ 20_T_請求明細 更新完了 %d行", len(invoices))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--mock", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    m = "mock" if args.mock else ("apply" if args.apply else "dry-run")
    setup_logging(LOG_DIR / f"sync_invoices_{datetime.now():%Y-%m-%d}.log")
    return run(m)


if __name__ == "__main__":
    sys.exit(main())
