"""MF取引先 → 10_M_取引先 の自動マッピング（rapidfuzz）。

Tier:
    ≥ 95%: 自動承認（mf_match_status = auto）
    80〜95%: 要確認キュー（queued）→ Teams通知
    < 80%: unmatched（手動/新規取引先候補）

実行:
    python sync_partners.py              # dry-run (MF API呼ぶがExcelは更新せず)
    python sync_partners.py --mock       # サンプルMFデータで動作確認
    python sync_partners.py --apply      # Excelを更新
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook
from rapidfuzz import fuzz

HERE = Path(__file__).parent
SCHEMA_PATH = HERE.parent / "sheet_builder" / "schema.yaml"
LOG_DIR = HERE / "logs"
SAMPLES_DIR = HERE / "samples"

AUTO_THRESHOLD = 95
QUEUE_THRESHOLD = 80

# RC名 → MF名 の手動エイリアス（英語↔日本語、略称↔正式名など）
PARTNER_ALIASES: dict[str, str] = {
    # normalize_name() 適用後のキー → MF側の検索文字列
    "oishiifarm": "oishii",
    "だらジャパン": "darajapan",
    "eukarya": "ユーカリヤ",          # Eukaryaは別途処理済みだが念のため
    "一般社団法人tabanone": "tabanone",
}

logger = logging.getLogger("sync_partners")


@dataclass
class PartnerMatch:
    partner_id: str
    partner_name: str
    mf_partner_id: str | None
    mf_match_score: int
    mf_match_status: str  # auto / queued / unmatched


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


_NORMALIZE_PATTERNS = [
    (re.compile(r"(株式会社|\(株\)|㈱)"), ""),
    (re.compile(r"(有限会社|\(有\)|㈲)"), ""),
    (re.compile(r"(合同会社|合資会社|合名会社)"), ""),
    (re.compile(r"(Inc\.?|LLC|Ltd\.?|Corp\.?|Co\.?)", re.IGNORECASE), ""),
    (re.compile(r"[\s　・,\.\-]"), ""),
]


def normalize_name(name: str) -> str:
    s = name or ""
    for pat, repl in _NORMALIZE_PATTERNS:
        s = pat.sub(repl, s)
    return s.lower()


def match_partners(rc_partners: list[dict], mf_partners: list[dict]) -> list[PartnerMatch]:
    """rc_partners: [{partner_id, partner_name}]; mf_partners: [{id, name}]"""
    mf_norm = [(m, normalize_name(m["name"])) for m in mf_partners]
    results: list[PartnerMatch] = []
    for rc in rc_partners:
        rc_norm = normalize_name(rc["partner_name"])
        # 手動エイリアスが定義されていればそちらを優先
        rc_norm_for_match = PARTNER_ALIASES.get(rc_norm, rc_norm)
        best_score = 0
        best_mf: dict | None = None
        for mf, mf_name_norm in mf_norm:
            score = fuzz.token_set_ratio(rc_norm_for_match, mf_name_norm)
            if score > best_score:
                best_score = score
                best_mf = mf
        if best_score >= AUTO_THRESHOLD:
            status = "auto"
        elif best_score >= QUEUE_THRESHOLD:
            status = "queued"
        else:
            status = "unmatched"
            best_mf = None
        results.append(PartnerMatch(
            partner_id=rc["partner_id"],
            partner_name=rc["partner_name"],
            mf_partner_id=(best_mf or {}).get("id"),
            mf_match_score=int(best_score),
            mf_match_status=status,
        ))
    return results


def load_mf_partners(mock: bool) -> list[dict]:
    if mock:
        f = SAMPLES_DIR / "mf_partners_samples.json"
        if not f.exists():
            logger.warning("サンプルがありません: %s", f)
            return []
        return json.loads(f.read_text(encoding="utf-8"))
    load_dotenv(HERE / ".env")
    from mf_client import MFClient, MFConfig
    client = MFClient(MFConfig.from_env())
    return list(client.list_partners())


def load_rc_partners_from_excel(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["10_M_取引先"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append({"partner_id": row[0], "partner_name": row[1] or ""})
    return rows


def write_matches_to_excel(xlsx_path: Path, matches: list[PartnerMatch]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb["10_M_取引先"]
    by_id = {m.partner_id: m for m in matches}
    for row_idx in range(2, ws.max_row + 1):
        pid = ws.cell(row=row_idx, column=1).value
        m = by_id.get(pid)
        if not m:
            continue
        ws.cell(row=row_idx, column=4, value=m.mf_partner_id or "")
        ws.cell(row=row_idx, column=5, value=m.mf_match_score)
        ws.cell(row=row_idx, column=6, value=m.mf_match_status)
    wb.save(xlsx_path)


def run(mode: str) -> int:
    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    xlsx = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]

    if not xlsx.exists():
        logger.error("Excelが存在しません: %s", xlsx)
        return 2

    rc_partners = load_rc_partners_from_excel(xlsx)
    mf_partners = load_mf_partners(mode == "mock")
    logger.info("RC取引先: %d件, MF取引先: %d件", len(rc_partners), len(mf_partners))

    matches = match_partners(rc_partners, mf_partners)
    tier = {"auto": 0, "queued": 0, "unmatched": 0}
    for m in matches:
        tier[m.mf_match_status] += 1
    logger.info("マッチ結果: auto=%d / queued=%d / unmatched=%d",
                tier["auto"], tier["queued"], tier["unmatched"])

    preview = [m.__dict__ for m in matches]
    (LOG_DIR / "partner_match_preview.json").write_text(
        json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    if mode in ("apply", "mock"):
        write_matches_to_excel(xlsx, matches)
        logger.info("✅ Excel更新完了 (%s)", mode)
    else:
        logger.info("[%s] Excelは更新せず。logs/partner_match_preview.json を確認", mode.upper())
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--mock", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    m = "mock" if args.mock else ("apply" if args.apply else "dry-run")
    setup_logging(LOG_DIR / f"sync_partners_{datetime.now():%Y-%m-%d}.log")
    return run(m)


if __name__ == "__main__":
    sys.exit(main())
