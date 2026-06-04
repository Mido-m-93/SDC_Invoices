"""統合管理Excel（RCP_契約請求支払統合管理.xlsx）を openpyxl で生成。

schema.yaml の定義に基づいて10シートを持つ雛形を作成する。
データは空のまま、ヘッダ・列幅・書式・凍結行だけ整える。

実行:
    python build_sheet.py            # sandbox/ に生成
    python build_sheet.py --prod     # 本番パスに生成（要承認）
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
SCHEMA_PATH = HERE / "schema.yaml"
LOG_DIR = HERE / "logs"

logger = logging.getLogger("build_sheet")


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def _header_style(fmt: dict) -> tuple[Font, PatternFill, Border, Alignment]:
    h = fmt["header"]
    font = Font(bold=h.get("bold", True), color="1F2937")
    fill = PatternFill("solid", fgColor=h.get("fill", "E8EEF7"))
    thin = Side(border_style=h.get("border", "thin"), color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, border, align


def _build_readme(ws, schema: dict) -> None:
    ws["A1"] = "RCP 契約請求支払 統合管理シート"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "目的: 営業→契約→請求→回収の一気通貫チェックを半自動化する",
        "構造: 裏正規化テーブル（M_*, T_*）+ 表横展開ビュー（V_*）",
        "",
        "【自動更新シート】",
        "  10_M_取引先       ← Notion Company + MF取引先（rapidfuzzで突合）",
        "  11_M_案件         ← Notion 🧪 Test_Unified_Pipeline v2 ミラー",
        "  12_M_契約         ← 契約スナップショット（契約フォルダ整備後に自動化予定）",
        "  20_T_請求明細     ← MoneyForward 請求書ミラー (read-only)",
        "  21_T_入金明細     ← MoneyForward 会計取引（入金）",
        "  22_T_支払明細     ← MoneyForward 会計取引（支払）",
        "  30_V_案件ビュー   ← 案件×契約×請求×入金 横展開",
        "  31_V_週次チェック ← 突合アラート（Teams投稿と同内容）",
        "  32_V_月次ダッシュボード ← 月次締めKPI",
        "",
        "【更新タイミング】",
        "  毎営業日 08:30 JST : Notion→M_* / MF→T_* 同期 + V_* 再計算 + Teams通知",
        "  毎月曜 09:00 JST   : 週次レポート",
        "  月初 09:00 JST     : 月次締めレポート",
        "",
        "【Safety】",
        "  - 本ファイルは自動生成の雛形。手編集した値は次回同期で上書きされる可能性があるため、",
        "    手コメントや手入力は 00_README / 12_M_契約.notes / 31_V_週次チェック.action 等に限定する",
        "  - MFとNotionは Source of Truth。差分があればそちら側を直す",
        "",
        f"  生成日時: {datetime.now():%Y-%m-%d %H:%M:%S}",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=text)
    ws.column_dimensions["A"].width = 100
    ws.freeze_panes = "A1"


def build(output_path: Path, schema: dict) -> None:
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    fmt = schema["formatting"]
    font, fill, border, align = _header_style(fmt)

    for sheet_def in schema["sheets"]:
        name = sheet_def["name"]
        ws = wb.create_sheet(title=name)

        if sheet_def["type"] == "doc":
            _build_readme(ws, schema)
            continue

        # ヘッダ行
        columns = sheet_def.get("columns", [])
        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx, value=col["name"])
            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = align
            ws.column_dimensions[get_column_letter(idx)].width = col.get("width", 14)
            # 列コメント（noteがあれば）
            if col.get("note"):
                cell.comment = None  # openpyxl Commentは重いので省略。noteはschema.yamlに残す

            # 書式（2行目以降のテンプレ書式だけ先頭行に設定、行追加時に継承される想定）
            fmt_code = col.get("fmt")
            if fmt_code:
                ws.cell(row=2, column=idx).number_format = fmt_code
            elif col.get("type") == "date":
                ws.cell(row=2, column=idx).number_format = fmt["date_fmt"]
            elif col.get("type") == "datetime":
                ws.cell(row=2, column=idx).number_format = fmt["datetime_fmt"]

        # 凍結
        ws.freeze_panes = sheet_def.get("freeze", "A2")

        # オートフィルタ
        if columns:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col}1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("✅ Excel生成完了: %s", output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="統合管理Excelの雛形を生成")
    parser.add_argument("--prod", action="store_true", help="本番パスに出力（要承認）")
    parser.add_argument("--output", type=Path, default=None, help="出力先を明示")
    args = parser.parse_args()

    setup_logging(LOG_DIR / f"build_sheet_{datetime.now():%Y-%m-%d}.log")

    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)

    if args.output:
        out = args.output
    elif args.prod:
        out = Path(schema["workbook"]["prod_dir"]) / schema["workbook"]["filename"]
        logger.warning("[PROD] 本番パスへ書き込みます: %s", out)
    else:
        out = Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]

    build(out, schema)
    return 0


if __name__ == "__main__":
    sys.exit(main())
