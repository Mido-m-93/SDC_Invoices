"""新Unified v2 → 11_M_案件 / 10_M_取引先 への同期（骨格）。

注意:
    現段階はスケルトン。新Unified v2 の data_source_id が config.yaml に埋まり、
    サンプルまたは本番データでmerge_pipelineが流れた後に機能する。

実行:
    python sync_from_notion.py           # sandbox 内の xlsx を更新
    python sync_from_notion.py --mock    # samples/*.json から擬似更新
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

HERE = Path(__file__).parent
SCHEMA_PATH = HERE / "schema.yaml"
NOTION_SYNC_DIR = HERE.parent / "notion_sync"
LOG_DIR = HERE / "logs"

logger = logging.getLogger("sync_from_notion")


def setup_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def load_records(mock: bool) -> list[dict]:
    """merge_pipeline の出力を入力にする。"""
    if mock:
        plan = NOTION_SYNC_DIR / "logs" / "merge_plan_preview.json"
        if not plan.exists():
            logger.warning("先に `python ../notion_sync/merge_pipeline.py --mock` を実行してください")
            return []
        return json.loads(plan.read_text(encoding="utf-8"))
    # 本番モード：将来ここで Notion API から直接新Unified v2を読む
    logger.info("（本番同期は未実装。現時点は --mock で動作確認）")
    return []


def make_partner_id(seen: dict[str, str], company: str) -> str:
    if company in seen:
        return seen[company]
    pid = f"RC-PRT-{len(seen) + 1:04d}"
    seen[company] = pid
    return pid


def sync(xlsx_path: Path, records: list[dict]) -> None:
    if not xlsx_path.exists():
        logger.error("Excelが存在しません。先に build_sheet.py を実行してください: %s", xlsx_path)
        return

    wb = load_workbook(xlsx_path)
    partners = wb["10_M_取引先"]
    projects = wb["11_M_案件"]

    # 既存行をクリア（ヘッダだけ残す）
    for ws in (partners, projects):
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    seen_partners: dict[str, str] = {}
    now_iso = datetime.now().isoformat(timespec="seconds")

    for i, r in enumerate(records, start=1):
        company = r.get("company") or ""
        pid = make_partner_id(seen_partners, company)
        projects.append([
            f"RC-PRJ-{i:04d}",
            r.get("project_name") or "",
            pid,
            r.get("status") or "",
            r.get("source_db") or "",
            r["properties"].get("Sales_In_Charge") or "",
            r["properties"].get("Region") or "",
            r["properties"].get("Project_Type") or "",
            r["properties"].get("New_Continuing") or "",
            r["properties"].get("Industry") or "",
            r["properties"].get("Probability") or "",
            r["properties"].get("Strategic_Fit") or "",
            r["properties"].get("Priority") or "",
            r["properties"].get("Expected_Close") or "",
            r["properties"].get("Last_Discussion") or "",
            r["properties"].get("Revenue_千円") or "",
            r["properties"].get("Revenue_月_千円") or "",
            r["properties"].get("PJ_Term_months") or "",
            r["properties"].get("Next_Step") or "",
            "",  # notion_url（後で埋める）
            r.get("original_url") or "",
            now_iso,
        ])

    # partners シート埋める
    for company, pid in seen_partners.items():
        partners.append([
            pid, company, company, "", "", "unmatched",
            "", "", "", "", "", ""
        ])

    wb.save(xlsx_path)
    logger.info("✅ 同期完了: %d件の案件, %d件の取引先 → %s", len(records), len(seen_partners), xlsx_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="新Unified v2 → 統合管理Excel 同期")
    parser.add_argument("--mock", action="store_true", help="merge_pipeline --mock の出力を使う")
    parser.add_argument("--xlsx", type=Path, default=None, help="対象Excelを明示")
    args = parser.parse_args()

    setup_logging(LOG_DIR / f"sync_from_notion_{datetime.now():%Y-%m-%d}.log")

    with SCHEMA_PATH.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)

    xlsx = args.xlsx or Path(schema["workbook"]["sandbox_dir"]) / schema["workbook"]["filename"]
    records = load_records(args.mock)
    sync(xlsx, records)
    return 0


if __name__ == "__main__":
    sys.exit(main())
